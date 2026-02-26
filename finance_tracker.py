import pandas as pd
import re
import os
import joblib
import logging
import argparse
from typing import Optional, Dict

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

class TransactionCategorizer:
    """Handles classification of transactions using ML or fallback keywords."""
    
    FALLBACK_KEYWORDS = {
        "Food": ["swiggy", "zomato", "mcdonald", "kfc", "restaurant", "food", "cafe", "starbucks", "burger", "bake", "pizza", "stall", "dine"],
        "Travel": ["uber", "ola", "train", "flight", "bus", "petrol", "auto", "shell", "hpcl", "metro", "toll", "parking", "fuel"],
        "Shopping": ["amazon", "flipkart", "myntra", "ajio", "zara", "h&m", "mall", "decathlon", "croma", "reliance digital", "apparel"],
        "Bills": ["electricity", "internet", "wifi", "broadband", "bill", "recharge", "mobile", "jio", "airtel", "tax", "maintenance"],
        "Subscriptions": ["netflix", "spotify", "hotstar", "prime", "youtube", "icloud", "gym", "software", "premium"],
        "Groceries": ["grocery", "dmart", "big bazaar", "supermarket", "blinkit", "instamart", "zepto", "bigbasket", "milk", "meat", "veg"],
        "Others": ["transfer", "cash", "misc", "pharmacy", "doctor", "gift", "salon", "cinema", "movie"]
    }

    def __init__(self, model_path: str = "expense_model.pkl", vectorizer_path: str = "expense_vectorizer.pkl", force_fallback: bool = False):
        self.use_ml = False
        self.model = None
        self.vectorizer = None

        if not force_fallback:
            self._load_ml_model(model_path, vectorizer_path)

    def _load_ml_model(self, model_path: str, vectorizer_path: str) -> None:
        try:
            self.model = joblib.load(model_path)
            self.vectorizer = joblib.load(vectorizer_path)
            self.use_ml = True
            logger.info("Machine Learning model loaded successfully.")
        except FileNotFoundError:
            logger.warning("ML model files not found. Falling back to keyword categorization.")

    def categorize(self, text: str) -> str:
        text_str = str(text).lower()
        clean_text = re.sub(r"[^a-z0-9\s]", " ", text_str).strip()
        
        # Determine Income vs Expense
        if "salary" in clean_text or "credit" in clean_text or "dividend" in clean_text or "refund" in clean_text:
             # Often these represent income, but we mainly check positive amounts later
             # We can classify them formally as Income here
             if "salary" in clean_text or "dividend" in clean_text:
                 if not self.use_ml:
                     return "Income"
        
        if self.use_ml and clean_text:
            text_vector = self.vectorizer.transform([clean_text])
            prediction = self.model.predict(text_vector)
            return prediction[0]
            
        # Fallback keyword logic
        for category, keywords in self.FALLBACK_KEYWORDS.items():
            for kw in keywords:
                if kw in clean_text:
                    return category
        return "Others"


class FinanceReportGenerator:
    """Processes financial CSV data and generates Styled Excel reports."""
    
    def __init__(self, categorizer: TransactionCategorizer):
        self.categorizer = categorizer
        self.df = pd.DataFrame()

    def process_csv(self, file_path: str) -> bool:
        """Loads and cleans the CSV."""
        logger.info(f"Loading data from {file_path}")
        try:
            self.df = pd.read_csv(file_path, encoding='utf-8')
        except FileNotFoundError:
            logger.error(f"Error: Could not find file {file_path}")
            return False
            
        self.df.columns = [c.strip() for c in self.df.columns]
        
        if "Description" not in self.df.columns or "Amount" not in self.df.columns:
            logger.error("CSV must include 'Description' and 'Amount' columns.")
            return False

        self._clean_data()
        self._apply_categorization()
        return True

    def _clean_data(self) -> None:
        """Cleans numerical formats without losing Income/Expense sign integrity."""
        self.df["Amount"] = (
            self.df["Amount"]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("₹", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        
        # Handle parentheses representing negative values `(500)` -> `-500`
        mask_parentheses = self.df["Amount"].str.startswith("(") & self.df["Amount"].str.endswith(")")
        self.df.loc[mask_parentheses, "Amount"] = "-" + self.df.loc[mask_parentheses, "Amount"].str[1:-1]
        
        self.df["Amount"] = pd.to_numeric(self.df["Amount"], errors="coerce").fillna(0.0)

    def _apply_categorization(self) -> None:
        self.df["Description"] = self.df["Description"].astype(str)
        self.df["Category"] = self.df["Description"].apply(self.categorizer.categorize)

    def export_excel(self, output_prefix: str = "finance_report") -> None:
        """Generates the styled Excel workbook."""
        if self.df.empty:
            logger.error("No data to export.")
            return

        logger.info("Generating Excel Report...")
        
        # Segregate data
        df_exp = self.df[self.df["Amount"] < 0].copy()
        df_exp["AbsAmount"] = df_exp["Amount"].abs()
        df_inc = self.df[self.df["Amount"] > 0].copy()
        
        summary = df_exp.groupby("Category")["AbsAmount"].sum().reset_index()
        total_expense = df_exp["AbsAmount"].sum()
        total_income = df_inc["Amount"].sum()

        wb = Workbook()
        self._build_transactions_sheet(wb.active, self.df)
        self._build_summary_sheet(wb.create_sheet("Summary"), summary, total_income, total_expense)

        # File Naming
        counter = 1
        final_name = f"{output_prefix}.xlsx"
        while os.path.exists(final_name):
            final_name = f"{output_prefix}_{counter}.xlsx"
            counter += 1

        wb.save(final_name)
        logger.info(f"Excel file created successfully: {final_name}")

        try:
            os.startfile(final_name)
        except Exception:
            pass

    def _build_transactions_sheet(self, ws, df):
        """Builds and styles the Raw Transactions sheet."""
        ws.title = "Transactions"
        
        # Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        
        headers = list(df.columns)
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Rows
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # Styling Income/Expense and Col Widths
        expense_fill = PatternFill("solid", fgColor="FFC7CE")
        income_fill = PatternFill("solid", fgColor="C6EFCE")
        
        amt_col_idx = headers.index("Amount") + 1
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=amt_col_idx)
            val = cell.value
            if isinstance(val, (int, float)):
                if val < 0:
                    cell.fill = expense_fill
                elif val > 0:
                    cell.fill = income_fill
                    
        self._auto_adjust_columns(ws)

    def _build_summary_sheet(self, ws, summary, total_inc, total_exp):
        """Builds the Summary Data, Overviews, and Pie Chart."""
        ws.title = "Summary"
        
        header_font = Font(bold=True)
        
        # High-level Overview
        ws.append(["Overview", "Amount"])
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        
        ws.append(["Total Income", total_inc])
        ws.append(["Total Expenses", total_exp])
        ws.append(["Net Balance", total_inc - total_exp])
        
        # Color coding overview
        ws["B2"].fill = PatternFill("solid", fgColor="C6EFCE") # Green Income
        ws["B3"].fill = PatternFill("solid", fgColor="FFC7CE") # Red Expense
        ws["B4"].font = Font(bold=True)

        ws.append([]) # Empty Row
        
        # Category Breakdown
        start_row = 6
        ws.append(["Category", "Total Spent"])
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=2).font = header_font
        
        for r in dataframe_to_rows(summary, index=False, header=False):
            ws.append(r)

        self._auto_adjust_columns(ws)

        if not summary.empty:
            pie = PieChart()
            pie.title = "Spending by Category"
            labels = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(summary))
            data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(summary))
            
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            ws.add_chart(pie, "D2")

    def _auto_adjust_columns(self, ws):
        """Auto scales column widths based on contents."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width


def main():
    parser = argparse.ArgumentParser(description="Automate Personal Finance tracking and Excel Generation.")
    parser.add_argument("input", help="Path to the input CSV file containing 'Description' and 'Amount' columns.")
    parser.add_argument("--output", "-o", default="finance_report", help="Prefix for the generated Excel file.")
    parser.add_argument("--no-ml", action="store_true", help="Force disable machine learning and run strictly on keyword fallback rules.")
    
    args = parser.parse_args()

    # Initialize Components
    categorizer = TransactionCategorizer(force_fallback=args.no_ml)
    generator = FinanceReportGenerator(categorizer)

    if generator.process_csv(args.input):
        generator.export_excel(args.output)


if __name__ == "__main__":
    main()